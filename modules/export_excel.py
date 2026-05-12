"""Excel出力処理です。"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_COLUMNS = [
    "発注日",
    "使用日",
    "食材名",
    "必要量",
    "発注単位",
    "発注数量",
    "仕入先",
    "備考",
    "OCR信頼度",
    "要確認フラグ",
]


def build_order_dataframe(candidates: pd.DataFrame, order_date: str, use_date: str) -> pd.DataFrame:
    """発注書Excelの列に合わせたデータを作ります。"""

    rows: list[dict[str, object]] = []
    order_targets = candidates[candidates["要確認"] != True] if "要確認" in candidates.columns else candidates
    for _, row in order_targets.iterrows():
        food_name = row.get("食材名", row.get("補正後食材名", ""))
        required_amount = row.get("必要量", row.get("数量", ""))
        order_quantity = row.get("発注数量", row.get("数量", ""))
        if not food_name or str(order_quantity).strip() in {"", "0", "数量要確認", "nan", "NaN"}:
            continue
        rows.append(
            {
                "発注日": order_date,
                "使用日": use_date,
                "食材名": food_name,
                "必要量": required_amount,
                "発注単位": row.get("発注単位", row.get("単位", "")) or row.get("単位", ""),
                "発注数量": order_quantity,
                "仕入先": row.get("仕入先", ""),
                "備考": row.get("備考", ""),
                "OCR信頼度": row.get("OCR信頼度", ""),
                "要確認フラグ": "",
            }
        )

    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


def dataframe_to_excel_bytes(order_df: pd.DataFrame) -> bytes:
    """Excelファイルのバイト列を作ります。

    MVPでは新規Excelを作ります。既存フォーマットへの転記は、実ファイル確認後に対応します。
    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "発注書"

    for row in dataframe_to_rows(order_df, index=False, header=True):
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="FDE9D9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in worksheet.iter_rows(min_row=2):
        flag_cell = row[OUTPUT_COLUMNS.index("要確認フラグ")]
        if flag_cell.value:
            for cell in row:
                cell.fill = warning_fill

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 40)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
